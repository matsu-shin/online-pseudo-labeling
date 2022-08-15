import openpyxl
import numpy as np
import json
from glob import glob
from PIL import Image
from tqdm import tqdm
import pickle

# 0~41:supervised image, 42~50:useless supervised image, 100~:un-supervised image
image_dict = {0: 'H1700285HE', 1: 'H16-13099_5_tumor_HE', 2: 'H1701984', 3: 'H02-09154_12_tumor_HE', 4: 'H1702766', 5: 'H1702480', 6: 'H1702105', 7: 'H02-09154_10-11_tumor_HE', 8: 'H17-09215_3_tumor_HE', 9: 'H17-09136_4_tumor_HE',
              10: 'H17-09192_10_tumor_HE', 11: 'H13-08405_11_tumor_HE', 12: 'H17-09624_3_tumor_HE', 13: 'H1702947_6', 14: 'H1700331', 15: 'H08-06230_9_tumor_HE', 16: 'H17-05649_3_tumor_HE', 17: 'H1701830', 18: 'H06-01625_4_tumor_HE',
              19: 'H17-09508_3_tumor_HE', 20: 'H1702553', 21: 'H1701147', 22: 'H1700062', 23: 'H17-09192_4_tumor_HE', 24: 'H17-07944_3_tumor_HE', 25: 'H01-07609_9_tumor_HE', 26: 'H1701428', 27: 'H07-02367_1_tumor_HE',
              28: 'H07-02244_12_tumor_HE', 29: 'H1702498', 30: 'H1700283', 31: 'H07-08889_12_tumor_HE', 32: 'H1700047', 33: 'H1701706', 34: 'H17-06307_3_tumor_HE', 35: 'H1700763', 36: 'H1702633',
              37: 'H17-09541_1_tumor_HE', 38: 'H1702528_4', 39: 'H02-09154_13_tumor_HE', 40: 'H1701140', 41: 'H17-08440_3_tumor_HE',
              42: 'H07-01097_4_tumor_HE', 43: 'H08-07550_10_tumor_HE', 44: 'H09-06091_5_tumor_HE', 45: 'H13-08405_8_tumor_HE', 46: 'H13-08405_9_tumor_HE', 47: 'H13-11840_1_tumor_HE', 48: 'H14-01231_1_tumor_HE', 49: 'H14-01231_3_tumor_HE', 50: 'H17-05498_1_HE',
              100: 'H14-08691_1_tumor_HE', 101: 'H09-07832_1_tumor_HE', 102: 'H06-03483_1_tumor_HE', 103: 'H10-01150_3_tumor_HE', 104: 'H05-05760_1_tumor_HE', 105: 'H13-11909_4_tumor_HE', 106: 'H10-08257_3_tumor_HE', 107: 'H11-10524_19_tumor_HE', 108: 'H13-03795_3_HE', 109: 'H13-10437_5_tumor_HE', 110: 'H13-01826_2_tumor_HE', 111: 'H05-09031_8_tumor_HE', 112: 'H13-09909_3_tumor_HE', 113: 'H05-09681_8-9_tumor_HE', 114: 'H13-00729_5_HE', 115: 'H03-03504_14_tumor_HE',
              116: 'H06-11200_1_tumor_HE', 117: 'H11-00615_7_tumor_HE', 118: 'H12-04875_4_tumor_HE', 119: 'H12-00339_9_tumor_HE', 120: 'H14-06549_2_tumor_HE', 121: 'H14-06593_2_tumor_HE', 122: 'H09-01437_8_tumor_HE', 123: 'H12-11417_3_tumor_HE', 124: 'H07-10227_1_tumor_HE', 125: 'H13-04401_5_tumor_HE', 126: 'H08-01323_7_tumor_HE', 127: 'H13-04622_2_tumor_HE', 128: 'H01-09953_6_tumor_HE', 129: 'H06-08736_2_tumor_HE', 130: 'H12-03674_3_tumor_HE', 131: 'H14-00706_3_tumor_HE',
              132: 'H14-09971_2_tumor_HE', 133: 'H10-03465_4_tumor_HE', 134: 'H14-11854_2_tumor_HE', 135: 'H08-02043_3_tumor_HE', 136: 'H11-04838_7_tumor_HE', 137: 'H06-06561_6_tumor_HE', 138: 'H12-03059_9_tumor_HE', 139: 'H11-07127_2_tumor_HE', 140: 'H13-02305_3_tumor_HE', 141: 'H09-05966_9_tumor_HE', 142: 'H08-07588_7_tumor_HE', 143: 'H08-06109_10_tumor_HE', 144: 'H09-04687_5_tumor_HE', 145: 'H03-08366_6-7_tumor_HE', 146: 'H14-03589_3_tumor_HE', 147: 'H05-03675_5_tumor_HE',
              148: 'H06-05983_6_tumor_HE', 149: 'H14-06356_7_tumor_HE',
              150: 'H01-07002_2_tumor_HE', 151: 'H14-09719_2_tumor_HE', 152: 'H10-05691_3_tumor_HE', 153: 'H10-03799_6_tumor_HE', 154: 'H10-07672_8_tumor_HE', 155: 'H10-07463_12_tumor_HE', 156: 'H09-01342_3_HE', 157: 'H07-02367_3_tumor_HE', 158: 'H11-04642_6_tumor_HE', 159: 'H12-07642_8_tumor_HE', 160: 'H14-02549_2_tumor_HE', 161: 'H13-02066_2_tumor_HE', 162: 'H14-08458_4_tumor_HE', 163: 'H09-03940_13_tumor_HE', 164: 'H11-03199_6_tumor_HE', 165: 'H13-00157_2_HE',
              166: 'H10-00160_4_tumor_HE', 167: 'H03-05031_12_tumor_HE', 168: 'H06-08863_2_tumor_HE', 169: 'H08-08190_2_tumor_HE', 170: 'H10-10859_2_tumor_HE', 171: 'H11-08954_6_tumor_HE', 172: 'H12-01122_4_tumor_HE', 173: 'H13-08783_4_HE', 174: 'H12-02317_8_tumor_HE', 175: 'H07-10854_11_tumor_HE', 176: 'H12-03987_6_tumor_HE', 177: 'H01-08092_15_tumor_HE', 178: 'H14-07631_9_tumor_HE', 179: 'H12-09129_4_tumor_HE', 180: 'H08-05796_14_tumor_HE', 181: 'H12-06101_3_tumor_HE',
              182: 'H04-06811_9_tumor_HE', 183: 'H12-08479_3', 184: 'H04-07017_7_tumor_HE', 185: 'H11-05247_5_tumor_HE', 186: 'H14-08147_2_tumor_HE', 187: 'H10-11208_7_tumor_HE', 188: 'H07-06502_3_tumor_HE', 189: 'H09-01211_5_HE', 190: 'H03-00970_3-4_tumor_HE', 191: 'H09-01143_7_HE', 192: 'H13-04883_5_tumor_HE', 193: 'H13-11538_2_tumor_HE', 194: 'H11-02680_14_tumor_HE', 195: 'H13-01216_2_tumor_HE', 196: 'H07-02244_11_tumor_HE', 197: 'H13-10325_2_tumor_HE',
              198: 'H14-01599_2_tumor_HE', 199: 'H13-07982_2_tumor_HE', 200: 'H03-10002_7-8_tumor_HE', 201: 'H13-06497_4_tumor_HE', 202: 'H14-09854_2_tumor_HE', 203: 'H11-11541_9', 204: 'H09-03029_3_tumor_HE', 205: 'H13-02458_3_HE', 206: 'H14-05307_7_tumor_HE', 207: 'H11-05400_3_tumor_HE', 208: 'H13-04034_5_tumor_HE', 209: 'H13-09458_3_HE', 210: 'H05-01019_11_tumor_HE', 211: 'H08-00126_9_tumor_HE', 212: 'H03-08498_7_tumor_HE', 213: 'H08-10462_4_tumor_HE',
              214: 'H13-07184_2_tumor_HE', 215: 'H06-07525_12_tumor_HE', 216: 'H08-07975_3_tumor_HE', 217: 'H08-00380_3-4_tumor_HE', 218: 'H05-10111_2_tumor_HE', 219: 'H11-09890_3', 220: 'H14-06311_1_tumor_HE', 221: 'H10-02625_8_tumor_HE', 222: 'H09-05070_2_tumor_HE', 223: 'H07-05630_3_tumor_HE', 224: 'H11-06574_11_tumor_HE', 225: 'H12-03065_1_tumor_HE', 226: 'H08-09185_9_tumor_HE', 227: 'H08-05737_6_tumor_HE', 228: 'H11-12368_5_tumor_HE', 229: 'H08-00683_4_tumor_HE',
              230: 'H12-06696_5_tumor_HE', 231: 'H13-04283_1_tumor_HE', 232: 'H08-09248_4_tumor_HE', 233: 'H09-08643_2_tumor_HE', 234: 'H14-10494_3_tumor_HE', 235: 'H12-01199_7_HE',
              236: 'H04-08460_4-5_tumor_HE', 237: 'H12-09770_7_tumor_HE', 238: 'H02-00567_3-4_tumor_HE', 239: 'H11-09653_5_tumor_HE', 240: 'H13-08439_1_tumor_HE', 241: 'H10-07520_9_tumor_HE', 242: 'H02-01100_8_tumor_HE', 243: 'H02-04313_1-3_tumor_HE', 244: 'H09-02628_3_tumor_HE', 245: 'H14-03497_2_tumor_HE', 246: 'H10-03742_6_tumor_HE', 247: 'H14-02049_3_tumor_HE', 248: 'H03-01946_8-9_tumor_HE', 249: 'H09-09440_12_tumor_HE', 250: 'H09-05734_4_tumor_HE', 251: 'H12-12223_4_tumor_HE',
              252: 'H06-08836_2_tumor_HE', 253: 'H09-05837_14_tumor_HE', 254: 'H13-04870_4_tumor_HE', 255: 'H05-01415_8_tumor_HE', 256: 'H13-02967_3_HE', 257: 'H10-01482_6_tumor_HE', 258: 'H11-06346_8_tumor_HE', 259: 'H13-11340_2_tumor_HE', 260: 'H14-11120_3_tumor_HE', 261: 'H01-09305_10_tumor_HE', 262: 'H10-09650_7_tumor_HE', 263: 'H11-06577_13_tumor_HE', 264: 'H11-03708_12_tumor_HE', 265: 'H10-01766_1_tumor_HE', 266: 'H08-04308_2_tumor_HE', 267: 'H11-00307_8_tumor_HE',
              268: 'H10-06167_9_tumor_HE',
              269: 'H14-00855_4_tumor_HE', 270: 'H09-07722_2_tumor_HE', 271: 'H14-11971_3_tumor_HE', 272: 'H12-12508_1_tumor_HE', 273: 'H02-04518_2-3_tumor_HE', 274: 'H12-07326_7_tumor_HE', 275: 'H11-08600_4_tumor_HE', 276: 'H04-10307_4-5_tumor_HE', 277: 'H05-05243_1_tumor_VBHE', 278: 'H10-08856_1_tumor_HE', 279: 'H10-09866_4_tumor_HE', 280: 'H12-10904_5_tumor_HE', 281: 'H02-05411_1-2_tumor_HE', 282: 'H14-06029_3_tumor_HE', 283: 'H14-11498_5_tumor_HE', 284: 'H11-10292_2_tumor_HE',
              285: 'H11-01390_6',
              286: 'H12-06421_1_tumor_HE', 287: 'H03-08850_4_tumor_HE', 288: 'H12-03080_6', 289: 'H06-09347_2_tumor_HE', 290: 'H05-06205_6_tumor_HE', 291: 'H10-09900_5_tumor_HE', 292: 'H09-07803_15_tumor_HE', 293: 'H06-00468_1_tumor_HE', 294: 'H11-05946_3_tumor_HE', 295: 'H06-07518_5_tumor_HE', 296: 'H02-08706_7_tumor_HE', 297: 'H12-04401_4_tumor_HE', 298: 'H06-03969_7-8_tumor_HE', 299: 'H13-10682_4_HE', 300: 'H08-09458_5-6_tumor_HE', 301: 'H11-02289_4_HE',
              302: 'H13-04255_5_tumor_HE',
              303: 'H07-02748_10_tumor_HE', 304: 'H05-08858_4_tumor_HE', 305: 'H13-08699_2_tumor_HE', 306: 'H02-00388_5-6_tumor_HE', 307: 'H14-04034_3_tumor_HE', 308: 'H05-01416_4_tumor_HE', 309: 'H08-04059_11_tumor_HE', 310: 'H12-03571_2_tumor_HE', 311: 'H11-06467_4_tumor_HE', 312: 'H10-04703_7-8_tumor_HE', 313: 'H14-10401_2_tumor_HE', 314: 'H08-06676_10_tumor_HE', 315: 'H04-05999_3_tumor_HE', 316: 'H14-10580_10_tumor_HE', 317: 'H09-05575_1_tumor_HE', 318: 'H09-01886_12_tumor_HE',
              319: 'H07-07733_20_tumor_HE',
              320: 'H11-04962_5_tumor_HE', 321: 'H05-00580_7_tumor_HE', 322: 'H12-01071_3_HE', 323: 'H10-11100_3_tumor_HE', 324: 'H07-09474_3_tumor_HE', 325: 'H11-08039_4_tumor_HE', 326: 'H09-00108_3_HE', 327: 'H07-09568_4_tumor_HE', 328: 'H11-04552_6_tumor_HE', 329: 'H10-04167_7_tumor_HE', 330: 'H11-02742_7_tumor_HE', 331: 'H03-00064_2-3_tumor_HE', 332: 'H10-08629_8_tumor_HE', 333: 'H07-05578_2_tumor_HE', 334: 'H14-06121_6_tumor_HE', 335: 'H12-08182_4_tumor_HE',
              336: 'H05-04707_6_tumor_HE',
              337: 'H07-07999_3_tumor_HE', 338: 'H06-00637_5_tumor_HE', 339: 'H05-06869_6_tumor_HE', 340: 'H09-06824_4_tumor_HE', 341: 'H11-05594_7', 342: 'H14-12029_7_tumor_HE', 343: 'H07-09511_6_tumor_HE', 344: 'H10-11717_4_tumor_HE', 345: 'H12-01311_5_tumor_HE', 346: 'H11-11173_3_tumor_HE', 347: 'H13-03954_1_tumor_HE', 348: 'H02-09766_11_tumor_HE', 349: 'H10-04023_10_tumor_HE', 350: 'H11-09811_9_tumor_HE', 351: 'H10-11875_3_tumor_HE', 352: 'H12-09667_5_tumor_HE',
              353: 'H14-10046_9_tumor_HE',
              354: 'H09-08890_5_tumor_HE', 355: 'H03-10164_6_tumor_HE', 356: 'H12-06934_5_tumor_HE', 357: 'H13-10706_5_tumor_HE', 358: 'H06-09222_2_tumor_HE', 359: 'H13-02239_2_tumor_HE', 360: 'H11-07184_5_tumor_HE', 361: 'H11-11317_4', 362: 'H10-08863_6_tumor_HE', 363: 'H03-07062_8-9_tumor_HE', 364: 'H12-11373_3_tumor_HE', 365: 'H07-10373_6_tumor_HE', 366: 'H13-09150_4_tumor_HE', 367: 'H13-02325_5_HE', 368: 'H10-07941_4_tumor_HE', 369: 'H08-07839_5_tumor_HE',
              370: 'H12-12527_8_tumor_HE',
              371: 'H13-02306_2_tumor_HE', 372: 'H05-03048_9-10_tumor_HE', 373: 'H12-07733_10', 374: 'H03-03055_2_tumor_HE', 375: 'H10-03896_3_tumor_HE', 376: 'H13-00384_4_tumor_HE', 377: 'H13-12346_2_HE', 378: 'H01-04626_2_tumor_HE', 379: 'H14-08419_4_tumor_HE', 380: 'H06-06744_1-4_tumor_HE', 381: 'H09-03626_3_tumor_HE', 382: 'H06-07279_9_tumor_HE', 383: 'H11-02150_1_tumor_HE', 384: 'H05-10241_2_tumor_HE', 385: 'H14-03310_4_tumor_HE', 386: 'H08-03917_6_tumor_HE',
              387: 'H07-02877_1-2_tumor_HE',
              388: 'H04-06617_6_tumor_HE', 389: 'H10-10593_5_tumor_HE', 390: 'H08-00622_2-3_tumor_HE', 391: 'H09-00020_5_HE', 392: 'H13-06441_1-2_tumor_HE', 393: 'H08-06288_1_tumor_HE', 394: 'H14-10156_2_tumor_HE', 395: 'H11-00204_5_tumor_HE', 396: 'H05-01817_4_tumor_HE', 397: 'H12-01905_12_tumor_HE', 398: 'H11-02033_7_tumor_HE', 399: 'H14-10581_2_tumor_HE', 400: 'H06-08947_8_tumor_HE', 401: 'H06-06830_3_tumor_HE', 402: 'H01-10522_4-5_tumor_HE', 403: 'H09-04350_3_tumor_HE',
              404: 'H11-06623_3_tumor_HE',
              405: 'H10-10387_2_tumor_HE', 406: 'H09-04548_7_tumor_HE', 407: 'H14-07387_1_tumor_HE', 408: 'H13-02618_7_tumor_HE', 409: 'H09-02092_2_tumor_HE', 410: 'H12-02316_1_tumor_HE', 411: 'H10-07285_12_tumor_HE', 412: 'H13-00119_3', 413: 'H12-02592_2_tumor_HE', 414: 'H14-01690_2_tumor_HE', 415: 'H04-08530_9_tumor_HE', 416: 'H14-05341_2_tumor_HE', 417: 'H06-10081_1_tumor_HE', 418: 'H03-08501_1-2_tumor_HE', 419: 'H06-11003_3_tumor_HE', 420: 'H14-02718_3_tumor_HE',
              421: 'H12-03427_4_tumor_HE',
              422: 'H11-02405_5_tumor_HE', 423: 'H10-11694_4_tumor_HE', 424: 'H07-07450_8-9_tumor_HE', 425: 'H08-07120_5_tumor_HE', 426: 'H10-05304_4_tumor_HE', 427: 'H13-04739_4_tumor_HE', 428: 'H11-02739_2_tumor_HE', 429: 'H14-09808_2_tumor_HE', 430: 'H05-08164_1-2_tumor_HE', 431: 'H13-07359_2_tumor_HE', 432: 'H12-00235_5_tumor_HE', 433: 'H02-09966_10_tumor_HE', 434: 'H10-07736_7_tumor_HE', 435: 'H14-09199_4_tumor_HE', 436: 'H09-01967_5_tumor_HE', 437: 'H13-10355_2_tumor_HE',
              438: 'H14-09591_1_tumor_HE', 439: 'H02-08705_8_tumor_HE', 440: 'H13-04931_5_tumor_HE', 441: 'H13-10866_3_tumor_HE', 442: 'H04-02215_10-11_tumor_HE', 443: 'H11-03147_7_tumor_HE', 444: 'H13-00278_4_tumor_HE', 445: 'H12-08046_5_tumor_HE', 446: 'H14-11103_3_tumor_HE', 447: 'H02-04092_4_tumor_HE', 448: 'H07-10555_9_tumor_HE', 449: 'H12-00336_7_tumor_HE', 450: 'H14-03856_4_tumor_HE', 451: 'H12-08257_5_tumor_HE', 452: 'H10-06052_13_tumor_HE', 453: 'H14-11733_3_tumor_HE',
              454: 'H09-05346_8_tumor_HE', 455: 'H06-05269_4_tumor_HE', 456: 'H04-04729_11_tumor_HE', 457: 'H04-06061_21_tumor_HE', 458: 'H13-07372_9_HE', 459: 'H12-09753_7_tumor_HE', 460: 'H14-06227_7_tumor_HE', 461: 'H04-08410_1_tumor_HE', 462: 'H10-03660_10_tumor_HE', 463: 'H10-06699_8_tumor_HE', 464: 'H11-00556_8_tumor_HE', 465: 'H08-04117_4_tumor_HE', 466: 'H09-04594_6_tumor_HE', 467: 'H12-08318_3_tumor_HE', 468: 'H07-03259_12_tumor_HE', 469: 'H10-10999_2_tumor_HE',
              470: 'H11-05687_4_tumor_HE',
              471: 'H08-07513_8_tumor_HE', 472: 'H12-03930_6_tumor_HE', 473: 'H12-00204_1_tumor_HE', 474: 'H11-01984_6_tumor_HE', 475: 'H03-03306_4-5_tumor_HE', 476: 'H09-02066_11_tumor_HE', 477: 'H11-00932_7-8_tumor_HE', 478: 'H11-03609_9_tumor_HE', 479: 'H11-08145_2_tumor_HE', 480: 'H06-10719_6_tumor_HE', 481: 'H11-10973_6_tumor_HE', 482: 'H13-05684_6_tumor_HE', 483: 'H11-04867_1_tumor_HE', 484: 'H11-08190_3_tumor_HE', 485: 'H02-07984_7_tumor_HE', 486: 'H14-03713_3_tumor_HE',
              487: 'H12-06456_5_tumor_HE', 488: 'H10-11490_10_tumor_HE', 489: 'H11-11456_5_tumor_HE', 490: 'H02-07289_5-7_tumor_HE', 491: 'H12-02517_3_tumor_HE', 492: 'H08-02259_5_tumor_HE', 493: 'H14-08883_4_tumor_HE', 494: 'H12-03309_9', 495: 'H10-02895_10_tumor_HE', 496: 'H05-07047_4_tumor_HE', 497: 'H04-07543_7_tumor_HE', 498: 'H14-05365_1_tumor_HE', 499: 'H12-00714_4_HE', 500: 'H03-04695_3-4_tumor_HE', 501: 'H10-11350_3_tumor_HE',
              502: 'H12-04478_6'}

# wb = openpyxl.load_workbook('../dataset/WSI/class_ratio_5class.xlsx')
# ws = wb["classratio"]

# proportion_dict = {}
# # idx = 100
# for idx in range(100, 503):
#     for rows in ws.iter_rows(min_row=2):
#         wsi_name = image_dict[idx]
#         if rows[0].value == wsi_name:
#             proportion = []
#             for i in range(1, 6):
#                 proportion.append(rows[i].value)
#             if len(glob('../dataset/WSI/unlabeled/'+wsi_name+'/_/*')) == 0:
#                 print('none patch data')
#             if len(glob('../dataset/WSI/unlabeled/'+wsi_name+'/_/*')) != 0 and sum(proportion) > 0.999:
#                 proportion_dict[idx] = proportion
# print(proportion_dict)


# with open("../dataset/WSI/image_name_dict.pkl", "wb") as tf:
#     pickle.dump(image_dict, tf)
# with open("../dataset/WSI/proportion_dict.pkl", "wb") as tf:
#     pickle.dump(proportion_dict, tf)

# with open("../dataset/WSI/image_name_dict.pkl", "rb") as tf:
#     image_name_dict = pickle.load(tf)
# print(image_name_dict[100])
# with open("../dataset/WSI/proportion_dict.pkl", "rb") as tf:
#     proportion_dict = pickle.load(tf)
# print(proportion_dict[100])

# # create train_bag_data
# with open("../dataset/WSI/image_name_dict.pkl", "rb") as tf:
#     image_name_dict = pickle.load(tf)

# train_bag_data = {}
# for idx in tqdm(range(100, 503)):
#     bag_data = []
#     wsi_name = image_name_dict[idx]
#     instance_path_list = glob('../dataset/WSI/unlabeled/'+wsi_name+'/_/*')
#     for path in instance_path_list:
#         data = Image.open(path)
#         data = np.asarray(data.convert('RGB'))
#         bag_data.append(np.array(data))
#     bag_data = np.array(bag_data)
#     train_bag_data[idx] = bag_data
# with open("../dataset/WSI/train_bag_data.pkl", "wb") as tf:
#     pickle.dump(train_bag_data, tf)
# with open("../dataset/WSI/train_bag_data.pkl", "rb") as tf:
#     train_bag_data = pickle.load(tf)
# print(train_bag_data[100].shape)


##########################################################################
# [:10]
# proportion_dict = {}
# # idx = 100
# for idx in range(100, 110):
#     for rows in ws.iter_rows(min_row=2):
#         if rows[0].value == image_dict[idx]:
#             proportion = []
#             for i in range(1, 6):
#                 proportion.append(rows[i].value)
#             if sum(proportion) > 0.999:
#                 proportion_dict[idx] = proportion
# print(proportion_dict)

# # not_used_idx = [124, 134, 261, 270, 353]
# # np.save('../dataset/WSI/image_name_dict.npy', image_dict)
# # np.save('../dataset/WSI/proportion_dict.npy', proportion_dict)

# with open("../dataset/WSI/image_name_dict_10.pkl", "wb") as tf:
#     pickle.dump(image_dict, tf)
# with open("../dataset/WSI/proportion_dict_10.pkl", "wb") as tf:
#     pickle.dump(proportion_dict, tf)

# with open("../dataset/WSI/image_name_dict_10.pkl", "rb") as tf:
#     image_name_dict = pickle.load(tf)
# print(image_name_dict[100])
# with open("../dataset/WSI/proportion_dict_10.pkl", "rb") as tf:
#     proportion_dict = pickle.load(tf)
# print(proportion_dict[100])

# # create train_bag_data
# with open("../dataset/WSI/image_name_dict_10.pkl", "rb") as tf:
#     image_name_dict = pickle.load(tf)

# train_bag_data = {}
# for idx in tqdm(range(100, 110)):
#     bag_data = []
#     wsi_name = image_name_dict[idx]
#     instance_path_list = glob('../dataset/WSI/unlabeled/'+wsi_name+'/_/*')
#     for path in instance_path_list:
#         data = Image.open(path)
#         data = np.asarray(data.convert('RGB'))
#         bag_data.append(np.array(data))
#     bag_data = np.array(bag_data)
#     train_bag_data[idx] = bag_data
# with open("../dataset/WSI/train_bag_data_10.pkl", "wb") as tf:
#     pickle.dump(train_bag_data, tf)
# with open("../dataset/WSI/train_bag_data_10.pkl", "rb") as tf:
#     train_bag_data = pickle.load(tf)
# print(train_bag_data[100].shape)


#########################################################################
# labeled data
dataset_path = '../dataset/WSI/labeled/'

wsi_path_list = glob(dataset_path+'/*')[:20]
print(wsi_path_list)

patch_path_list = []
for path in wsi_path_list:
    patch_path_list.extend(glob(path+'/_/*'))
print(len(patch_path_list))

data_list, label_list = [], []
for path in tqdm(patch_path_list):
    data = Image.open(path)
    data = np.asarray(data.convert('RGB'))
    label = int(path[-5: -4])
    data_list.append(data)
    label_list.append(label)
data = np.array(data_list)
label = np.array(label_list)
print(data.shape, label.shape)
np.save('../dataset/WSI/test_data.npy', data)
np.save('../dataset/WSI/test_label.npy', label)
